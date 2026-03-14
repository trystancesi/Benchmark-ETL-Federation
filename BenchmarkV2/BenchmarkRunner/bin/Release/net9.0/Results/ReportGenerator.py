"""
ReportGenerator.py  v4  (source : raw_*.csv)
Graphiques generes avec matplotlib (PNG) et inseres dans l'Excel.

Onglets :
  1. Résumé exécutif
  2-5. T1 a T4  (barres groupées par volume)
  6. T5 - Burst SELECT  (montée en charge)
  7. T6 - Ingestion ETL
  8. T7 - Burst INSERT  (montée en charge)
  9. Données

Usage : python ReportGenerator.py [dossier_results]
"""

import sys, os, glob, io
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Palette ───────────────────────────────────────────────────────────────────
C_HDR_BG = "1F3864"
C_HDR_FG = "FFFFFF"
C_SUB_BG = "D6E4F7"
C_ALT    = "F2F7FC"
C_BORDER = "BDD7EE"

ARCH_COLORS = {
    "ETL":                   "#2E75B6",
    "FederatedLinkedServer": "#ED7D31",
    "FederatedInMemory":     "#70AD47",
}
ARCH_MARKERS = {
    "ETL":                   "o",
    "FederatedLinkedServer": "s",
    "FederatedInMemory":     "^",
}
ARCH_LABELS = {
    "ETL":                   "ETL (centralise)",
    "FederatedLinkedServer": "Federe - Linked Server",
    "FederatedInMemory":     "Federe - Mediation C#",
}
SCALE_ORDER = ["Small", "Medium", "Large"]
ARCHS       = ["ETL", "FederatedLinkedServer", "FederatedInMemory"]

thin   = Side(style="thin", color=C_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Helpers Excel ─────────────────────────────────────────────────────────────

def hdr(ws, row, col, value, bg=C_HDR_BG, fg=C_HDR_FG, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c

def cel(ws, row, col, value, bold=False, color="000000", bg=None,
        fmt=None, center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    c.border    = BORDER
    if bg:  c.fill = PatternFill("solid", fgColor=bg)
    if fmt: c.number_format = fmt
    return c

def title_row(ws, text, span, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def alt(i): return C_ALT if i % 2 == 0 else "FFFFFF"

def insert_png(ws, fig, anchor, width_px=700):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    buf.seek(0)
    img = XLImage(buf)
    scale = width_px / img.width
    img.width  = int(img.width  * scale)
    img.height = int(img.height * scale)
    ws.add_image(img, anchor)
    plt.close(fig)

# ── Matplotlib style global ───────────────────────────────────────────────────

def setup_mpl():
    plt.rcParams.update({
        "font.family":       "DejaVu Sans",
        "font.size":         10,
        "axes.titlesize":    12,
        "axes.titleweight":  "bold",
        "axes.labelsize":    10,
        "axes.spines.top":   False,
        "axes.spines.right": False,
        "axes.grid":         True,
        "grid.alpha":        0.35,
        "grid.linestyle":    "--",
        "legend.fontsize":   9,
        "legend.framealpha": 0.85,
        "figure.facecolor":  "white",
    })

# ── Graphiques matplotlib ─────────────────────────────────────────────────────

def fig_bar_standard(agg, scenario_label):
    scales = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    x = np.arange(len(scales))
    w = 0.25
    fig, ax = plt.subplots(figsize=(8, 5))
    for i, arch in enumerate(ARCHS):
        sub  = agg[agg["Architecture"] == arch]
        vals = [float(sub[sub["Scale"]==s]["MedianMs"].iloc[0])
                if not sub[sub["Scale"]==s].empty else 0 for s in scales]
        errs = [float(sub[sub["Scale"]==s]["IQR"].iloc[0]) / 2
                if not sub[sub["Scale"]==s].empty else 0 for s in scales]
        bars = ax.bar(x + i*w, vals, w, label=ARCH_LABELS[arch],
                      color=ARCH_COLORS[arch], alpha=0.88,
                      yerr=errs, error_kw={"capsize":4,"elinewidth":1.2,"capthick":1.2})
        for bar, v in zip(bars, vals):
            if v > 0:
                ax.text(bar.get_x() + bar.get_width()/2,
                        bar.get_height() + max(vals)*0.01,
                        f"{v:.0f}", ha="center", va="bottom", fontsize=8)
    ax.set_xticks(x + w); ax.set_xticklabels(scales)
    ax.set_ylabel("Mediane end-to-end (ms)")
    ax.set_title(f"{scenario_label} - Mediane par volume\n(barres d'erreur = IQR)")
    ax.legend(loc="upper left")
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:.0f}"))
    fig.tight_layout()
    return fig


def fig_metrics_standard(agg, scenario_label):
    scales = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    x = np.arange(len(scales))
    w = 0.25
    colors = [ARCH_COLORS[a] for a in ARCHS]

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle(f"{scenario_label} — Métriques détaillées", fontweight="bold", fontsize=13)

    specs = [
        (axes[0,0], "CpuPct",   "CPU moyen (%)",           "CPU (%)",     ".1f"),
        (axes[0,1], "MemSQL",   "Mémoire SQL Server (Mo)",  "Mo",          ".0f"),
        (axes[1,0], "LogReads", "Lectures logiques (médiane)", "lectures", ".0f"),
        (axes[1,1], None,       "Temps serveur vs End-to-end (ms)", "ms",  ".0f"),
    ]

    for ax, metric, title, ylabel, fmt in specs:
        if metric is not None:
            for i, arch in enumerate(ARCHS):
                sub  = agg[agg["Architecture"] == arch]
                vals = [float(sub[sub["Scale"]==s][metric].iloc[0])
                        if not sub[sub["Scale"]==s].empty else 0 for s in scales]
                bars = ax.bar(x + i*w, vals, w, label=ARCH_LABELS[arch],
                              color=colors[i], alpha=0.85)
                for bar, v in zip(bars, vals):
                    if v > 0:
                        ax.text(bar.get_x() + bar.get_width()/2,
                                bar.get_height() + max(vals)*0.01,
                                f"{v:{fmt}}", ha="center", va="bottom", fontsize=7)
            ax.set_xticks(x + w); ax.set_xticklabels(scales)
        else:
            ax.set_visible(False)

        ax.set_title(title, fontsize=10, fontweight="bold")
        ax.set_ylabel(ylabel)
        ax.legend(fontsize=7, loc="upper left")
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))

    fig.tight_layout()
    return fig


def fig_burst_metrics(agg, scale, scenario_label):
    workers = sorted(agg["ConcurrentWorkers"].unique())
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))
    fig.suptitle(f"{scenario_label} — Ressources système [{scale}]",
                 fontweight="bold", fontsize=12)

    specs = [
        (axes[0], "CpuPct",  "CPU moyen (%)",            "%"),
        (axes[1], "MemSQL",  "Mémoire SQL Server (Mo)",   "Mo"),
        (axes[2], None,      "Mémoire .NET (Mo)",         "Mo"),
    ]

    for ax, metric, title, ylabel in specs:
        col = "MemSQL" if metric == "MemSQL" else metric
        for arch in ARCHS:
            sub  = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
            if metric is None:
                vals = [float(sub[sub["ConcurrentWorkers"]==w]["MemDotNet"].iloc[0])
                        if not sub[sub["ConcurrentWorkers"]==w].empty and "MemDotNet" in sub.columns
                        else np.nan for w in workers]
            else:
                vals = [float(sub[sub["ConcurrentWorkers"]==w][col].iloc[0])
                        if not sub[sub["ConcurrentWorkers"]==w].empty else np.nan
                        for w in workers]
            ax.plot(workers, vals, color=ARCH_COLORS[arch],
                    marker=ARCH_MARKERS[arch], linewidth=2.2, markersize=7,
                    label=ARCH_LABELS[arch])
        ax.set_xticks(workers)
        ax.set_xlabel("Workers simultanés")
        ax.set_ylabel(ylabel)
        ax.set_title(title, fontsize=10, fontweight="bold")
        ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))

    fig.tight_layout()
    return fig


def fig_e2e_serveur(agg, scenario_label):
    scales = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    fig, axes = plt.subplots(1, 3, figsize=(14, 5), sharey=False)
    fig.suptitle(f"{scenario_label} — Temps serveur vs End-to-end\n"
                 f"(plein = E2E applicatif, hachuré = exécution serveur seul)",
                 fontweight="bold", fontsize=11)
    for i, arch in enumerate(ARCHS):
        ax  = axes[i]
        sub = agg[agg["Architecture"] == arch]
        e2e = [float(sub[sub["Scale"]==s]["MedianMs"].iloc[0])
               if not sub[sub["Scale"]==s].empty else 0 for s in scales]
        srv = [float(sub[sub["Scale"]==s]["ServerMs"].iloc[0])
               if not sub[sub["Scale"]==s].empty else 0 for s in scales]
        xi  = np.arange(len(scales))
        bw  = 0.35
        ax.bar(xi - bw/2, e2e, bw, label="E2E",
               color=ARCH_COLORS[arch], alpha=0.88)
        ax.bar(xi + bw/2, srv, bw, label="Serveur",
               color=ARCH_COLORS[arch], alpha=0.38, hatch="///")
        for j, (ve, vs) in enumerate(zip(e2e, srv)):
            if ve > 0:
                ax.text(j - bw/2, ve + max(e2e)*0.03,
                        f"{ve:.0f}", ha="center", va="bottom", fontsize=8)
            if vs > 0:
                ax.text(j + bw/2, vs + max(e2e)*0.03,
                        f"{vs:.0f}", ha="center", va="bottom", fontsize=8)
        ax.set_xticks(xi); ax.set_xticklabels(scales)
        ax.set_title(ARCH_LABELS[arch], fontsize=10, fontweight="bold",
                     color=ARCH_COLORS[arch])
        ax.set_ylabel("ms")
        ax.legend(fontsize=8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(axis="y", alpha=0.3, linestyle="--")
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    fig.tight_layout()
    return fig


def fig_burst_rps(agg, scale, scenario_label):
    workers = sorted(agg["ConcurrentWorkers"].unique())
    fig, ax = plt.subplots(figsize=(8, 5))
    for arch in ARCHS:
        sub  = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
        vals = []
        for w in workers:
            r = sub[sub["ConcurrentWorkers"] == w]
            vals.append(float(r["MedianRPS"].iloc[0]) if not r.empty else np.nan)
        ax.plot(workers, vals, color=ARCH_COLORS[arch],
                marker=ARCH_MARKERS[arch], linewidth=2.2, markersize=7,
                label=ARCH_LABELS[arch])
    ax.set_xticks(workers)
    ax.set_xlabel("Workers simultanees")
    ax.set_ylabel("Debit median (req/s)")
    ax.set_title(f"{scenario_label} - Debit [{scale}]\n(req/s = requetes par seconde, plus haut = meilleur)")
    ax.legend()
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:.0f}"))
    fig.tight_layout()
    return fig


def fig_burst_rps_combined(agg, scenario_label):
    """
    Graphique combiné débit — 3 subplots côte à côte (Small / Medium / Large)
    Exporté en haute résolution pour insertion dans le mémoire.
    """
    scales  = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    workers = sorted(agg["ConcurrentWorkers"].unique())

    fig, axes = plt.subplots(1, len(scales), figsize=(18, 5), sharey=False)
    fig.suptitle(f"{scenario_label} — Débit par volume\n"
                 f"(req/s = requêtes par seconde, plus haut = meilleur)",
                 fontweight="bold", fontsize=13)

    if len(scales) == 1:
        axes = [axes]

    for ax, scale in zip(axes, scales):
        for arch in ARCHS:
            sub  = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
            vals = [float(sub[sub["ConcurrentWorkers"]==w]["MedianRPS"].iloc[0])
                    if not sub[sub["ConcurrentWorkers"]==w].empty else np.nan
                    for w in workers]
            ax.plot(workers, vals,
                    color=ARCH_COLORS[arch],
                    marker=ARCH_MARKERS[arch],
                    linewidth=2.2, markersize=7,
                    label=ARCH_LABELS[arch])
        ax.set_xticks(workers)
        ax.set_xlabel("Workers simultanés")
        ax.set_ylabel("Débit médian (req/s)")
        ax.set_title(f"[{scale}]", fontsize=11, fontweight="bold")
        ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:.0f}"))
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(axis="y", alpha=0.3, linestyle="--")

    fig.tight_layout()
    return fig


def fig_burst_completion_combined(agg, scenario_label):
    """
    Graphique combiné temps de complétion E2E — 3 subplots côte à côte (Small / Medium / Large)
    Exporté en haute résolution pour insertion dans le mémoire.
    """
    scales  = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    workers = sorted(agg["ConcurrentWorkers"].unique())

    fig, axes = plt.subplots(1, len(scales), figsize=(18, 5), sharey=False)
    fig.suptitle(f"{scenario_label} — Temps de complétion E2E par volume\n"
                 f"(ms = millisecondes, bande = IQR, plus bas = meilleur)",
                 fontweight="bold", fontsize=13)

    if len(scales) == 1:
        axes = [axes]

    for ax, scale in zip(axes, scales):
        for arch in ARCHS:
            sub  = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
            vals, p25, p75 = [], [], []
            for w in workers:
                r = sub[sub["ConcurrentWorkers"] == w]
                if not r.empty:
                    vals.append(float(r["MedianMs"].iloc[0]))
                    p25.append(float(r["P25"].iloc[0]))
                    p75.append(float(r["P75"].iloc[0]))
                else:
                    vals.append(np.nan); p25.append(np.nan); p75.append(np.nan)
            ax.plot(workers, vals,
                    color=ARCH_COLORS[arch],
                    marker=ARCH_MARKERS[arch],
                    linewidth=2.2, markersize=7,
                    label=ARCH_LABELS[arch])
            ax.fill_between(workers, p25, p75,
                            color=ARCH_COLORS[arch], alpha=0.12)
        ax.set_xticks(workers)
        ax.set_xlabel("Workers simultanés")
        ax.set_ylabel("Temps de complétion médian (ms)")
        ax.set_title(f"[{scale}]", fontsize=11, fontweight="bold")
        ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:.0f}"))
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(axis="y", alpha=0.3, linestyle="--")

    fig.tight_layout()
    return fig


def fig_burst_memsql_combined(agg, scenario_label):
    """
    Graphique combiné mémoire SQL Server — 3 subplots côte à côte (Small / Medium / Large)
    Exporté en haute résolution pour insertion dans le mémoire.
    """
    scales  = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    workers = sorted(agg["ConcurrentWorkers"].unique())

    fig, axes = plt.subplots(1, len(scales), figsize=(18, 5), sharey=False)
    fig.suptitle(f"{scenario_label} — Mémoire SQL Server par volume\n"
                 f"(Mo = mégaoctets, plus bas = meilleur)",
                 fontweight="bold", fontsize=13)

    if len(scales) == 1:
        axes = [axes]

    for ax, scale in zip(axes, scales):
        for arch in ARCHS:
            sub  = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
            vals = [float(sub[sub["ConcurrentWorkers"]==w]["MemSQL"].iloc[0])
                    if not sub[sub["ConcurrentWorkers"]==w].empty else np.nan
                    for w in workers]
            ax.plot(workers, vals,
                    color=ARCH_COLORS[arch],
                    marker=ARCH_MARKERS[arch],
                    linewidth=2.2, markersize=7,
                    label=ARCH_LABELS[arch])
        ax.set_xticks(workers)
        ax.set_xlabel("Workers simultanés")
        ax.set_ylabel("Mémoire SQL Server (Mo)")
        ax.set_title(f"[{scale}]", fontsize=11, fontweight="bold")
        ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(axis="y", alpha=0.3, linestyle="--")

    fig.tight_layout()
    return fig


def fig_burst_latency(agg, scale, scenario_label):
    workers = sorted(agg["ConcurrentWorkers"].unique())
    fig, ax = plt.subplots(figsize=(8, 5))
    for arch in ARCHS:
        sub  = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
        vals, p25, p75 = [], [], []
        for w in workers:
            r = sub[sub["ConcurrentWorkers"] == w]
            if not r.empty:
                vals.append(float(r["MedianMs"].iloc[0]))
                p25.append(float(r["P25"].iloc[0]))
                p75.append(float(r["P75"].iloc[0]))
            else:
                vals.append(np.nan); p25.append(np.nan); p75.append(np.nan)
        ax.plot(workers, vals, color=ARCH_COLORS[arch],
                marker=ARCH_MARKERS[arch], linewidth=2.2, markersize=7,
                label=ARCH_LABELS[arch])
        ax.fill_between(workers, p25, p75, color=ARCH_COLORS[arch], alpha=0.12)
    ax.set_xticks(workers)
    ax.set_xlabel("Workers simultanés")
    ax.set_ylabel("Temps de complétion E2E (ms)")
    ax.set_title(f"{scenario_label} - Temps de complétion [{scale}]\n(ms = millisecondes, bande = IQR, plus bas = meilleur)")
    ax.legend()
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:.0f}"))
    fig.tight_layout()
    return fig


def fig_t6_bar(agg):
    scales   = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    vals     = [float(agg[agg["Scale"]==s]["RPS"].iloc[0]) for s in scales]
    durees_s = [float(agg[agg["Scale"]==s]["MedianMs"].iloc[0]) / 1000 for s in scales]
    fig, axes = plt.subplots(1, 2, figsize=(10, 4))

    bars = axes[0].bar(scales, vals, color=ARCH_COLORS["ETL"], alpha=0.88, width=0.45)
    for bar, v in zip(bars, vals):
        axes[0].text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(vals)*0.01,
                f"{v:,.0f}", ha="center", va="bottom", fontsize=9)
    axes[0].set_ylabel("Lignes / seconde")
    axes[0].set_title("Debit d'ingestion (lignes/s)")
    axes[0].yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))

    bars2 = axes[1].bar(scales, durees_s, color="#5B9BD5", alpha=0.88, width=0.45)
    for bar, v in zip(bars2, durees_s):
        axes[1].text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(durees_s)*0.01,
                f"{v:.1f}s", ha="center", va="bottom", fontsize=9)
    axes[1].set_ylabel("Duree mediane (s)")
    axes[1].set_title("Duree d'ingestion (secondes)")
    axes[1].yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{v:.0f}s"))

    fig.suptitle("T6 - Ingestion batch ETL (SqlBulkCopy)", fontweight="bold")
    fig.tight_layout()
    return fig

# ── Chargement & agrégation ───────────────────────────────────────────────────

def load_raw(results_dir):
    files = glob.glob(os.path.join(results_dir, "raw_*.csv"))
    if not files:
        print(f"Aucun fichier raw_*.csv trouve dans {results_dir}"); sys.exit(1)
    dfs = []
    for f in files:
        df = pd.read_csv(f)
        df.columns = [c.strip() for c in df.columns]
        dfs.append(df)
    df = pd.concat(dfs, ignore_index=True)
    df = df[df["IsWarmup"] == False]
    df = df[df["ErrorMessage"].isna() | (df["ErrorMessage"] == "")]
    df["Scale"] = df["Scale"].str.capitalize()
    return df


def agg_standard(df, scenario):
    sub = df[df["Scenario"] == scenario]
    if sub.empty: return pd.DataFrame()
    def stats(g):
        ms = g["EndToEndMs"]
        return pd.Series({
            "MedianMs": np.median(ms),
            "P25":      np.percentile(ms, 25),
            "P75":      np.percentile(ms, 75),
            "IQR":      np.percentile(ms, 75) - np.percentile(ms, 25),
            "CpuPct":   g["CpuPercent"].median(),
            "MemDotNet":g["MemoryMb"].median(),
            "MemSQL":   g["MemSqlServerMb"].median(),
            "LogReads": g["LogicalReads"].median(),
            "ServerMs": g["ServerElapsedMs"].median(),
        })
    return sub.groupby(["Architecture","Scale"]).apply(
        stats, include_groups=False).reset_index()


def agg_burst(df, scenario):
    sub = df[df["Scenario"] == scenario]
    if sub.empty: return pd.DataFrame()
    sub = sub.copy()
    sub["ConcurrentWorkers"] = sub["ConcurrentWorkers"].astype(int)
    def stats(g):
        ms = g["EndToEndMs"]
        return pd.Series({
            "MedianMs":  np.median(ms),
            "P25":       np.percentile(ms, 25),
            "P75":       np.percentile(ms, 75),
            "IQR":       np.percentile(ms, 75) - np.percentile(ms, 25),
            "MedianRPS": g["ThroughputRequestsPerSec"].median(),
            "Errors":    g["ErrorCount"].median(),
            "CpuPct":    g["CpuPercent"].median(),
            "MemSQL":    g["MemSqlServerMb"].median(),
            "MemDotNet": g["MemoryMb"].median(),
        })
    return sub.groupby(["Architecture","Scale","ConcurrentWorkers"]).apply(
        stats, include_groups=False).reset_index()


def agg_t6(df):
    sub = df[df["Scenario"] == "T6_ETLBatchIngestion"]
    if sub.empty: return pd.DataFrame()
    def stats(g):
        ms = g["EndToEndMs"]
        return pd.Series({
            "MedianMs":   np.median(ms),
            "P25":        np.percentile(ms, 25),
            "P75":        np.percentile(ms, 75),
            "IQR":        np.percentile(ms, 75) - np.percentile(ms, 25),
            "RPS":        g["IngestionRowsPerSecond"].median(),
            "RowsCopied": g["RowsCopied"].median(),
        })
    return sub.groupby(["Scale"]).apply(stats, include_groups=False).reset_index()

# ── Onglets Excel ─────────────────────────────────────────────────────────────

def build_standard_sheet(wb, agg, scenario_label):
    if agg.empty: return
    scales = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    ws = wb.create_sheet(scenario_label[:31])
    ws.sheet_view.showGridLines = False
    title_row(ws, scenario_label, 2 + len(scales) * 2)

    row = 3
    hdr(ws, row, 1, "Architecture")
    col = 2
    for sc in scales:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        hdr(ws, row, col, sc, bg=C_SUB_BG, fg=C_HDR_BG)
        col += 2
    row += 1
    hdr(ws, row, 1, "")
    col = 2
    for _ in scales:
        hdr(ws, row, col,   "Mediane (ms)", bg=C_SUB_BG, fg="000000", bold=False)
        hdr(ws, row, col+1, "IQR (ms)",     bg=C_SUB_BG, fg="000000", bold=False)
        col += 2
    row += 1

    for arch in ARCHS:
        sub = agg[agg["Architecture"] == arch]
        if sub.empty: continue
        bg = alt(row)
        cel(ws, row, 1, ARCH_LABELS[arch], bold=True,
            color=ARCH_COLORS[arch].lstrip("#"), bg=bg)
        col = 2
        for sc in scales:
            r   = sub[sub["Scale"] == sc]
            med = float(r["MedianMs"].iloc[0]) if not r.empty else None
            iqr = float(r["IQR"].iloc[0])      if not r.empty else None
            cel(ws, row, col,   round(med,1) if med else "—", bg=bg, fmt="0.0", center=True)
            cel(ws, row, col+1, round(iqr,1) if iqr else "—", bg=bg, fmt="0.0", center=True)
            col += 2
        row += 1

    set_widths(ws, [28] + [13, 10] * len(scales))
    fig = fig_bar_standard(agg, scenario_label)
    insert_png(ws, fig, f"A{row+2}", width_px=680)
    fig2 = fig_metrics_standard(agg, scenario_label)
    insert_png(ws, fig2, f"A{row+22}", width_px=900)
    fig3 = fig_e2e_serveur(agg, scenario_label)
    insert_png(ws, fig3, f"A{row+52}", width_px=900)


def build_burst_sheet(wb, agg, sheet_label, scenario_label):
    if agg.empty: return
    scales  = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    workers = sorted(agg["ConcurrentWorkers"].unique())

    ws = wb.create_sheet(sheet_label[:31])
    ws.sheet_view.showGridLines = False
    title_row(ws, sheet_label, 1 + len(workers) + 2)

    for si, scale in enumerate(scales):
        base_row = 3 + si * (len(ARCHS) + 4)
        ws.merge_cells(start_row=base_row, start_column=1,
                       end_row=base_row, end_column=len(workers)+2)
        c = ws.cell(row=base_row, column=1, value=f"Volume : {scale}")
        c.font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
        c.fill      = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[base_row].height = 22

        hdr(ws, base_row+1, 1, "Architecture")
        for j, w in enumerate(workers):
            hdr(ws, base_row+1, j+2, f"{int(w)} workers\n(ms)")
        hdr(ws, base_row+1, len(workers)+2, "10->200w\n(completion)")

        for i, arch in enumerate(ARCHS):
            row = base_row + 2 + i
            sub = agg[(agg["Architecture"] == arch) & (agg["Scale"] == scale)]
            bg  = alt(i)
            cel(ws, row, 1, ARCH_LABELS[arch], bold=True,
                color=ARCH_COLORS[arch].lstrip("#"), bg=bg)
            vals = []
            for j, w in enumerate(workers):
                r = sub[sub["ConcurrentWorkers"] == w]
                v = float(r["MedianMs"].iloc[0]) if not r.empty else None
                cel(ws, row, j+2, round(v,1) if v else "—", bg=bg, fmt="0.0", center=True)
                if v: vals.append(v)
            if len(vals) >= 2:
                evol  = (vals[-1] - vals[0]) / vals[0] * 100
                color = "C00000" if evol > 200 else ("70AD47" if evol < 50 else "000000")
                cel(ws, row, len(workers)+2,
                    f"+{evol:.0f}%" if evol >= 0 else f"{evol:.0f}%",
                    bg=bg, bold=True, color=color, center=True)

    set_widths(ws, [28] + [10]*len(workers) + [10])

    img_row   = 3 + len(scales) * (len(ARCHS) + 4) + 3
    col_right = get_column_letter(len(workers) + 5)

    # Graphiques combinés (nouveaux) — un seul graphique par métrique pour tous les volumes
    fig_rps_combined = fig_burst_rps_combined(agg, scenario_label)
    insert_png(ws, fig_rps_combined, f"A{img_row}", width_px=1100)

    fig_completion_combined = fig_burst_completion_combined(agg, scenario_label)
    insert_png(ws, fig_completion_combined, f"A{img_row+22}", width_px=1100)

    fig_memsql_combined = fig_burst_memsql_combined(agg, scenario_label)
    insert_png(ws, fig_memsql_combined, f"A{img_row+44}", width_px=1100)

    # Graphiques individuels par volume (conservés pour l'Excel)
    for si, scale in enumerate(scales):
        anchor_row = img_row + 66 + si * 52
        fig_rps = fig_burst_rps(agg, scale, scenario_label)
        insert_png(ws, fig_rps, f"A{anchor_row}", width_px=560)
        fig_lat = fig_burst_latency(agg, scale, scenario_label)
        insert_png(ws, fig_lat, f"{col_right}{anchor_row}", width_px=560)
        fig_met = fig_burst_metrics(agg, scale, scenario_label)
        insert_png(ws, fig_met, f"A{anchor_row+20}", width_px=1050)


def build_t6_sheet(wb, agg):
    if agg.empty: return
    ws = wb.create_sheet("T6 - Ingestion ETL")
    ws.sheet_view.showGridLines = False
    title_row(ws, "T6 - Ingestion batch ETL (SqlBulkCopy)", 5)

    row = 3
    for c, h in enumerate(["Volume","Duree mediane (s)","IQR (s)",
                            "Debit (lignes/s)","Lignes copiees"], 1):
        hdr(ws, row, c, h)
    row += 1

    scales = [s for s in SCALE_ORDER if s in agg["Scale"].unique()]
    for sc in scales:
        bg = alt(row)
        r  = agg[agg["Scale"] == sc]
        cel(ws, row, 1, sc, bold=True, bg=bg)
        cel(ws, row, 2, round(float(r["MedianMs"].iloc[0]) / 1000, 1) if not r.empty else "—",
            bg=bg, fmt="0.0", center=True)
        cel(ws, row, 3, round(float(r["IQR"].iloc[0]) / 1000, 1) if not r.empty else "—",
            bg=bg, fmt="0.0", center=True)
        cel(ws, row, 4, round(float(r["RPS"].iloc[0]),0) if not r.empty else "—",
            bg=bg, fmt="#,##0", center=True)
        cel(ws, row, 5, int(r["RowsCopied"].iloc[0]) if not r.empty else "—",
            bg=bg, fmt="#,##0", center=True)
        row += 1

    set_widths(ws, [14, 22, 14, 18, 18])
    row += 1
    ws.cell(row=row, column=1,
            value="Note : T6 exclusif ETL - pas d'equivalent federe.").font = \
        Font(name="Arial", italic=True, color="666666", size=9)

    fig = fig_t6_bar(agg)
    insert_png(ws, fig, f"A{row+2}", width_px=480)


def build_summary_sheet(wb, df_raw):
    ws = wb.create_sheet("Resume executif", 0)
    ws.sheet_view.showGridLines = False
    scales = [s for s in SCALE_ORDER if s in df_raw["Scale"].unique()]

    ws.merge_cells(f"A1:{get_column_letter(3+len(scales))}1")
    c = ws["A1"]
    c.value     = "Benchmark ETL vs Architecture Federee - Resume executif"
    c.font      = Font(name="Arial", bold=True, size=14, color=C_HDR_FG)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(3+len(scales))}2")
    ws["A2"].value     = "Mediane (ms) par scenario, architecture et volume - plus bas = meilleur"
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    hdr(ws, row, 1, "Scenario"); hdr(ws, row, 2, "Architecture")
    for j, sc in enumerate(scales):
        hdr(ws, row, 3+j, f"{sc} (ms)")
    if len(scales) > 1:
        hdr(ws, row, 3+len(scales), "Small->Large")
    row += 1

    scenarios = [
        ("T1_GetSNList",            "T1 - Liste SN"),
        ("T2_GetMesuresByPN",       "T2 - Jointures"),
        ("T3_GetMesuresHistorique", "T3 - Historique CTE"),
        ("T4_InsertMesure",         "T4 - INSERT mesure"),
    ]
    for sc_key, sc_label in scenarios:
        sub = df_raw[df_raw["Scenario"] == sc_key]
        if sub.empty: continue
        first = True
        for arch in ARCHS:
            a  = sub[sub["Architecture"] == arch]
            bg = alt(row)
            cel(ws, row, 1, sc_label if first else "", bold=first, bg=bg)
            cel(ws, row, 2, ARCH_LABELS[arch], bold=True,
                color=ARCH_COLORS[arch].lstrip("#"), bg=bg)
            vals = []
            for j, sc in enumerate(scales):
                v_arr = a[a["Scale"] == sc]["EndToEndMs"]
                v = float(np.median(v_arr)) if not v_arr.empty else None
                cel(ws, row, 3+j, round(v,1) if v else "—",
                    bg=bg, fmt="0.0", center=True)
                if v: vals.append(v)
            if len(scales) > 1 and len(vals) >= 2:
                evol  = (vals[-1] - vals[0]) / vals[0] * 100
                color = "C00000" if evol > 100 else ("70AD47" if evol < 20 else "000000")
                cel(ws, row, 3+len(scales),
                    f"+{evol:.0f}%" if evol >= 0 else f"{evol:.0f}%",
                    bg=bg, color=color, bold=True, center=True)
            first = False; row += 1
        row += 1

    for sc_key, label in [("T5_BurstSelect", "T5 - Burst SELECT (10w)"),
                           ("T7_BurstInsert", "T7 - Burst INSERT (10w)")]:
        sub = df_raw[(df_raw["Scenario"] == sc_key) & (df_raw["ConcurrentWorkers"] == 10)]
        if sub.empty: continue
        first = True
        for arch in ARCHS:
            a  = sub[sub["Architecture"] == arch]
            bg = alt(row)
            cel(ws, row, 1, label if first else "", bold=first, bg=bg)
            cel(ws, row, 2, ARCH_LABELS[arch], bold=True,
                color=ARCH_COLORS[arch].lstrip("#"), bg=bg)
            for j, sc in enumerate(scales):
                v_arr = a[a["Scale"] == sc]["ThroughputRequestsPerSec"]
                v = float(np.median(v_arr)) if not v_arr.empty else None
                cel(ws, row, 3+j, f"{v:.0f} rps" if v else "—", bg=bg, center=True)
            first = False; row += 1
        row += 1

    set_widths(ws, [24, 28] + [14]*len(scales) + [13])
    ws.freeze_panes = "A5"


def build_data_sheet(wb, df_raw):
    ws = wb.create_sheet("Donnees")
    ws.sheet_view.showGridLines = False
    agg_cols = ["Architecture","Scenario","Scale","ConcurrentWorkers"]
    metrics  = ["EndToEndMs","ServerElapsedMs","ServerCpuMs","CpuPercent",
                "MemoryMb","MemSqlServerMb","LogicalReads","PhysicalReads",
                "ThroughputRequestsPerSec","IngestionRowsPerSecond","ErrorCount"]
    rows = []
    for keys, g in df_raw.groupby(agg_cols, dropna=False):
        row_d = dict(zip(agg_cols, keys))
        for m in metrics:
            if m in g.columns:
                vals = g[m].dropna()
                row_d[m+"_median"] = round(np.median(vals),2) if len(vals) else None
        rows.append(row_d)
    out  = pd.DataFrame(rows)
    cols = list(out.columns)
    for j, c in enumerate(cols, 1):
        hdr(ws, 1, j, c)
    for i, (_, r) in enumerate(out.iterrows(), 2):
        bg = alt(i)
        for j, c in enumerate(cols, 1):
            v = r[c]
            cel(ws, i, j, None if pd.isna(v) else v, bg=bg)
    set_widths(ws, [20]*len(cols))

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    results_dir = sys.argv[1] if len(sys.argv) > 1 else "Results"
    if not os.path.isdir(results_dir):
        print(f"Dossier introuvable : {results_dir}"); sys.exit(1)

    print(f"Lecture des raw CSV dans {results_dir}...")
    df = load_raw(results_dir)
    print(f"  {len(df)} mesures - scales: {sorted(df['Scale'].unique())} - "
          f"scenarios: {sorted(df['Scenario'].unique())}")

    setup_mpl()
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, df)
    print("  OK Resume executif")

    for sc_key, sc_label in [
        ("T1_GetSNList",            "T1 - Liste SN (SELECT simple)"),
        ("T2_GetMesuresByPN",       "T2 - Mesures par piece (jointures)"),
        ("T3_GetMesuresHistorique", "T3 - Historique complet (CTE)"),
        ("T4_InsertMesure",         "T4 - INSERT mesure unitaire"),
    ]:
        build_standard_sheet(wb, agg_standard(df, sc_key), sc_label)
        print(f"  OK {sc_label}")

    build_burst_sheet(wb, agg_burst(df, "T5_BurstSelect"),
                      "T5 - Burst SELECT", "T5 - Burst SELECT")
    print("  OK T5 - Burst SELECT")

    build_t6_sheet(wb, agg_t6(df))
    print("  OK T6 - Ingestion ETL")

    build_burst_sheet(wb, agg_burst(df, "T7_BurstInsert"),
                      "T7 - Burst INSERT", "T7 - Burst INSERT")
    print("  OK T7 - Burst INSERT")

    build_data_sheet(wb, df)
    print("  OK Donnees")

    out = os.path.join(results_dir, "rapport_benchmark.xlsx")
    wb.save(out)
    print(f"\nRapport genere : {out}")


if __name__ == "__main__":
    main()